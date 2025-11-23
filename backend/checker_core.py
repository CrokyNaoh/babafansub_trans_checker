"""
翻译检查核心模块 - 保留完整 Excel 格式并支持颜色标注
基于 CHTchecker_v3.0.py 改造，支持多项目配置
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import os

os.environ["OPENPYXL_LXML"] = "False"


class ExcelChecker:
    """Excel 检查器 - 保留格式版本"""
    
    def __init__(self, err_dict, term_dict):
        """
        初始化检查器
        
        Args:
            err_dict: 错误词典 (包含 err, warn, repeat, transhint)
            term_dict: 术语词典 (包含 word 列表)
        """
        self.err_dict = err_dict
        self.term_dict = term_dict
    
    def fix_content(self, content):
        """清理内容"""
        return content.replace("\r\n", "").replace("\n", "").replace("　", "")
    
    def estimate_width(self, hints):
        """估算列宽"""
        if not hints:
            return 0
        return int(2 * max(map(lambda e: len(e), hints)))
    
    def estimate_height(self, hints):
        """估算行高"""
        return 20 * len(hints)
    
    def do_common_check(self, wb, sheetnames, input_col, err_col, warn_col):
        """
        检查易错字词模式
        
        Args:
            wb: openpyxl 工作簿
            sheetnames: 要检查的工作表名列表
            input_col: 输入列 (如 'C')
            err_col: 错误提示列 (如 'E')
            warn_col: 警告提示列 (如 'F')
        """
        pos = ord(input_col) - ord('A')
        err_pos = ord(err_col) - ord('A')
        warn_pos = ord(warn_col) - ord('A')
        
        err_words = self.err_dict.get('err', {})
        warn_words = self.err_dict.get('warn', {})
        repeat_chars = self.err_dict.get('repeat', [])
        
        for sheetname in sheetnames:
            sheet = wb[sheetname]
            first = True
            
            for row in sheet.iter_rows():
                if first:
                    first = False
                    continue
                
                if len(row) < (pos + 1) or row[pos].value is None:
                    continue
                
                content = str(row[pos].value)
                hint_err = []
                hint_warn = []
                hint_repeat = []
                mark_chars = []  # 需要标红的字符
                
                # 1. 检查错误词
                for err_word, correct_word in err_words.items():
                    if err_word in content:
                        hint_err.append(f"{err_word}→{correct_word}")
                
                # 2. 检查易混淆字（需要标红）
                for warn_word, desc in warn_words.items():
                    for ch in warn_word:
                        if ch in content:
                            mark_chars.append(ch)
                            hint_warn.append(f"{warn_word}：{desc}")
                            break
                
                # 3. 检查叠字
                for i in range(1, len(content)):
                    if content[i] in repeat_chars:
                        continue
                    if content[i] in hint_repeat:
                        continue
                    if content[i-1] == content[i]:
                        hint_repeat.append(content[i])
                
                if hint_repeat:
                    hint_warn.append(" ".join(["出現疊字"] + hint_repeat))
                
                # 设置错误和警告提示
                err_hint_text = "\r\n".join(hint_err)
                warn_hint_text = "\r\n".join(hint_warn)
                
                sheet.cell(row[0].row, err_pos + 1).value = err_hint_text
                sheet.cell(row[0].row, err_pos + 1).alignment = Alignment(wrapText=True)
                sheet.cell(row[0].row, warn_pos + 1).value = warn_hint_text
                sheet.cell(row[0].row, warn_pos + 1).alignment = Alignment(wrapText=True)
                
                # 调整列宽
                cd_err = sheet.column_dimensions[err_col]
                cd_err.width = max(cd_err.width or 0, self.estimate_width(hint_err))
                
                cd_warn = sheet.column_dimensions[warn_col]
                cd_warn.width = max(cd_warn.width or 0, self.estimate_width(hint_warn))
                
                # 调整行高
                rd = sheet.row_dimensions[row[0].row]
                if rd.height is None:
                    rd.height = 20
                rd.height = max(
                    rd.height,
                    self.estimate_height(hint_err),
                    self.estimate_height(hint_warn)
                )
                
                # 关键功能：对混淆字符标红（Rich Text）
                if mark_chars:
                    red_font = InlineFont(color='FF0000')
                    rich_text = CellRichText([
                        TextBlock(red_font, char) if char in mark_chars else char
                        for char in content
                    ])
                    sheet[input_col + str(row[0].row)] = rich_text
    
    def do_spec_check(self, wb, sheetnames, input_col, warn_col, include_transhint=False):
        """
        标注统一词模式
        
        Args:
            wb: openpyxl 工作簿
            sheetnames: 要检查的工作表名列表
            input_col: 输入列 (如 'C')
            warn_col: 警告提示列 (如 'E')
            include_transhint: 是否包含易塞翻词提示
        """
        pos = ord(input_col) - ord('A')
        warn_pos = ord(warn_col) - ord('A')
        
        term_words = self.term_dict.get('word', [])
        transhint_words = self.err_dict.get('transhint', {})
        
        for sheetname in sheetnames:
            sheet = wb[sheetname]
            first = True
            
            for row in sheet.iter_rows():
                if first:
                    first = False
                    continue
                
                if len(row) < (pos + 1) or row[pos].value is None:
                    continue
                
                content = str(row[pos].value)
                content_clean = self.fix_content(content)
                
                hint = []
                extra_hint = []
                
                # 1. 检查统一词
                for term in term_words:
                    term_text = term['ja']
                    term_text_clean = self.fix_content(term_text)
                    
                    if term_text_clean in content_clean:
                        comment = f"({term['note']})" if term.get('note') else ''
                        hint.append(f"{term_text}{comment}：{term['zh']}")
                
                # 2. 检查易塞翻词（可选）
                if include_transhint:
                    term_set = set(term['ja'] for term in term_words)
                    for trans_word, trans_desc in transhint_words.items():
                        if trans_word not in term_set and trans_word in content_clean:
                            extra_hint.append(f"{trans_word}：{trans_desc}")
                    
                    if extra_hint:
                        hint.append("【易塞翻詞，提供參考釋義】")
                        hint.extend(extra_hint)
                
                # 设置警告提示
                warn_hint_text = "\r\n".join(hint)
                sheet.cell(row[0].row, warn_pos + 1).value = warn_hint_text
                sheet.cell(row[0].row, warn_pos + 1).alignment = Alignment(wrapText=True)
                
                # 调整列宽和行高
                cd_warn = sheet.column_dimensions[warn_col]
                cd_warn.width = max(cd_warn.width or 0, self.estimate_width(hint))
                
                rd = sheet.row_dimensions[row[0].row]
                if rd.height is None:
                    rd.height = 20
                rd.height = max(rd.height, self.estimate_height(hint))
    
    def process_file(self, input_path, output_path, config):
        """
        处理文件主入口
        
        Args:
            input_path: 输入文件路径
            output_path: 输出文件路径
            config: 配置字典，包含：
                - mode: 'common' 或 'spec'
                - inputCol: 输入列
                - outputCol1: 输出列1（错误/警告）
                - outputCol2: 输出列2（仅 common 模式）
                - checkAllSheets: 是否检查所有工作表
                - includeTransHint: 是否包含易塞翻词（仅 spec 模式）
        
        Returns:
            输出文件路径
        """
        # 加载工作簿
        wb = load_workbook(input_path)
        
        # 确定要处理的工作表
        if config.get('checkAllSheets'):
            sheetnames = wb.sheetnames
        else:
            sheetnames = [wb.sheetnames[0]]
        
        # 执行检查
        if config['mode'] == 'common':
            self.do_common_check(
                wb,
                sheetnames,
                config['inputCol'],
                config['outputCol1'],
                config['outputCol2']
            )
        else:  # spec
            self.do_spec_check(
                wb,
                sheetnames,
                config['inputCol'],
                config['outputCol1'],
                config.get('includeTransHint', False)
            )
        
        # 保存（所有原始格式保留）
        wb.save(output_path)
        return output_path


